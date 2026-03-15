"""Generador XLSX multi-hoja con colores y formato forense — Multi-RMM."""
from __future__ import annotations
from pathlib import Path
from datetime import datetime
from ..models.base import RMMType, RMMSession, RMMConnection, ParseResult, ConnectionDirection
from ..models.enrichment import IPEnrichment
from ..models.incident import IncidentContext
from ..models.summary import ForensicSummary
from ..utils import fmt_dt, fmt_dur, clean, cprint

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
try:
    from colorama import Fore
except ImportError:
    class Fore:
        GREEN = YELLOW = RED = ""

# ─── Paleta de colores ────────────────────────────────────────────────────

TEAL_HEX   = "1A8080"
TEAL_LIGHT  = "E8F5F5"
WHITE       = "FFFFFF"
RED_HEX     = "C0392B"
ORANGE_HEX  = "D68910"
GREEN_HEX   = "1A7A47"

RISK_BG = {"CRÍTICO": "FDE8E8", "ALTO": "FEF5E4", "MEDIO": "FEFAE4", "BAJO": "E8F5EE"}
RISK_FG = {"CRÍTICO": RED_HEX, "ALTO": ORANGE_HEX, "MEDIO": "9A7A10", "BAJO": GREEN_HEX}

# Colores por RMM para la columna "RMM"
RMM_COLORS = {
    "AnyDesk":                {"bg": "1A8080", "fg": WHITE},
    "TeamViewer":             {"bg": "004680", "fg": WHITE},
    "ScreenConnect":          {"bg": "5B2C6F", "fg": WHITE},
    "Chrome Remote Desktop":  {"bg": "1A73E8", "fg": WHITE},
    "Splashtop":              {"bg": "E8600A", "fg": WHITE},
    "RustDesk":               {"bg": "2E7D32", "fg": WHITE},
}

PROXIMITY_BG = {
    "CRÍTICO (±24h)": "FDE8E8",
    "ALTO (±3d)":     "FEF5E4",
    "MEDIO (±7d)":    "FEFAE4",
}
PROXIMITY_FG = {
    "CRÍTICO (±24h)": RED_HEX,
    "ALTO (±3d)":     ORANGE_HEX,
    "MEDIO (±7d)":    "9A7A10",
}


# ─── Helpers de celdas ───────────────────────────────────────────────────

def _hdr(ws, row, col, val, bg=TEAL_HEX, fg=WHITE):
    c = ws.cell(row=row, column=col, value=clean(val))
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=True, color=fg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = Border(bottom=Side(border_style="thin", color="FFFFFF"))
    return c


def _cell(ws, row, col, val, bold=False, fg=None, bg=None, wrap=True):
    c = ws.cell(row=row, column=col, value=clean(val) if isinstance(val, str) else val)
    c.font = Font(bold=bold, color=fg or "1A2C2C")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(vertical="top", wrap_text=wrap)
    return c


def _rmm_cell(ws, row, col, rmm_name: str):
    """Escribe el nombre del RMM con color de fondo por RMM."""
    colors = RMM_COLORS.get(rmm_name, {"bg": "555555", "fg": WHITE})
    return _cell(ws, row, col, rmm_name, bold=True, fg=colors["fg"], bg=colors["bg"])


def _risk_cell(ws, row, col, risk: str):
    """Aplica formato de riesgo a una celda."""
    c = _cell(ws, row, col, risk, bold=True, fg=RISK_FG.get(risk, GREEN_HEX))
    c.fill = PatternFill("solid", fgColor=RISK_BG.get(risk, "E8F5EE"))
    return c


def _proximity_cell(ws, row, col, label: str):
    """Aplica formato de proximidad a una celda."""
    if label in PROXIMITY_BG:
        c = _cell(ws, row, col, label, bold=True, fg=PROXIMITY_FG.get(label, "1A2C2C"))
        c.fill = PatternFill("solid", fgColor=PROXIMITY_BG[label])
    else:
        c = _cell(ws, row, col, label or "—")
    return c


def _set_widths(ws, widths: list[int]):
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w


# ─── Generador principal ─────────────────────────────────────────────────

def generate_xlsx(
    summary: ForensicSummary,
    ip_results: list[IPEnrichment],
    outpath: str,
    incident: IncidentContext | None = None,
) -> None:
    if not HAS_OPENPYXL:
        cprint("  [!] openpyxl no disponible. Omitiendo XLSX.", Fore.YELLOW)
        return

    wb = openpyxl.Workbook()

    has_incident = incident is not None and incident.has_incident_date
    has_country = incident is not None and incident.has_country
    multiple_rmms = len(summary.results_by_rmm) > 1

    sessions = summary.all_sessions
    connections = summary.all_connections
    enriched_map: dict[str, IPEnrichment] = {e.ip: e for e in ip_results}

    # ══════════════════════════════════════════════════════════════════════
    #  Hoja 1: Resumen
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Resumen"
    _hdr(ws, 1, 1, "INFORME FORENSE MULTI-RMM — RESUMEN EJECUTIVO", bg="0F5555", fg=WHITE)
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 28

    rows_data: list[tuple[str, str | int]] = [
        ("METADATOS", ""),
        ("Fecha de generación", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("RMMs analizados", ", ".join(summary.rmm_types_found)),
        ("Hostnames encontrados", ", ".join(summary.hostnames_found) or "—"),
        ("Total archivos analizados",
         sum(len(pr.source_files) for pr in summary.results_by_rmm.values())),
    ]

    if summary.date_range_conn:
        rows_data.append(("Período de conexiones",
                          f"{fmt_dt(summary.date_range_conn[0])} — {fmt_dt(summary.date_range_conn[1])}"))
    if summary.date_range_sessions:
        rows_data.append(("Período de sesiones",
                          f"{fmt_dt(summary.date_range_sessions[0])} — {fmt_dt(summary.date_range_sessions[1])}"))

    if has_incident:
        rows_data.append(("Fecha de incidente", fmt_dt(incident.incident_date)))
    if has_country:
        rows_data.append(("País de origen", incident.origin_country))

    # Per-RMM breakdown
    rows_data += [("", ""), ("DESGLOSE POR RMM", "")]
    for rmm_name, pr in sorted(summary.results_by_rmm.items()):
        n_sess = len(pr.sessions)
        n_conn = len(pr.connections)
        n_ft = sum(s.file_transfers for s in pr.sessions)
        rows_data.append((
            f"  {rmm_name}",
            f"{n_sess} sesiones, {n_conn} conexiones, {n_ft} transferencias"
        ))

    # Global stats
    rows_data += [
        ("", ""),
        ("TOTALES GLOBALES", ""),
        ("Total conexiones", summary.total_connections),
        ("Conexiones entrantes", summary.incoming),
        ("Conexiones salientes", summary.outgoing),
        ("Total sesiones", summary.total_sessions),
        ("IDs remotos únicos", ", ".join(summary.unique_ids)),
        ("Total transferencias de archivos", summary.total_file_transfers),
        ("Total eventos de clipboard", summary.total_clipboard),
        ("Total transferencias de texto", summary.total_text_transfers),
        ("Sesiones con elevación de privilegios", summary.elevated_sessions),
        ("Máx archivos en clipboard simultáneos", summary.max_clipboard_files),
        ("IPs públicas únicas", len(summary.all_public_ips)),
    ]

    if has_incident:
        rows_data += [
            ("", ""),
            ("ANÁLISIS DE INCIDENTE", ""),
            ("Sesiones en ventana ±24h", len(summary.sessions_within_24h)),
            ("Sesiones en ventana ±3d", len(summary.sessions_within_3d)),
            ("Sesiones en ventana ±7d", len(summary.sessions_within_7d)),
        ]
        if has_country:
            rows_data += [
                ("Sesiones informativas (mismo país)", summary.informative_sessions),
                ("Sesiones sospechosas (otro país)", summary.suspicious_sessions),
            ]

    if multiple_rmms and summary.cross_rmm_ips:
        cross_count = sum(1 for rmms in summary.cross_rmm_ips.values() if len(rmms) > 1)
        rows_data += [
            ("", ""),
            ("CORRELACIÓN CROSS-RMM", ""),
            ("IPs compartidas entre RMMs", cross_count),
        ]

    for i, (k, v) in enumerate(rows_data, 2):
        c1 = ws.cell(row=i, column=1, value=clean(k))
        c2 = ws.cell(row=i, column=2, value=clean(str(v)) if not isinstance(v, (int, float)) else v)
        if not v and k:
            # Section header
            c1.font = Font(bold=True, color=TEAL_HEX)
            c1.fill = PatternFill("solid", fgColor=TEAL_LIGHT)
            ws.merge_cells(f"A{i}:B{i}")
        else:
            c1.font = Font(bold=bool(not v or k == k.upper()))

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 58

    # ══════════════════════════════════════════════════════════════════════
    #  Hoja 2: Sesiones
    # ══════════════════════════════════════════════════════════════════════
    ws_s = wb.create_sheet("Sesiones")
    hdrs_s = [
        "#", "RMM", "Hostname", "ID Remoto", "Alias", "Inicio", "Fin",
        "Duración", "Transferencias", "Clipboard", "Texto",
        "Elevación", "IPs", "Riesgo", "Score",
    ]
    if has_incident:
        hdrs_s.append("Proximidad Incidente")
    if has_country:
        hdrs_s.append("Clasificación País")

    for c, h in enumerate(hdrs_s, 1):
        _hdr(ws_s, 1, c, h)

    for i, s in enumerate(sessions, 2):
        col = 1
        _cell(ws_s, i, col, i - 1); col += 1
        _rmm_cell(ws_s, i, col, s.rmm_type.value); col += 1
        _cell(ws_s, i, col, s.hostname); col += 1
        _cell(ws_s, i, col, s.remote_id); col += 1
        _cell(ws_s, i, col, s.alias); col += 1
        _cell(ws_s, i, col, fmt_dt(s.start_dt)); col += 1
        _cell(ws_s, i, col, fmt_dt(s.end_dt)); col += 1
        _cell(ws_s, i, col, fmt_dur(s.duration_sec)); col += 1

        # Transferencias de archivos
        c_ft = _cell(ws_s, i, col, s.file_transfers); col += 1
        if s.file_transfers > 0:
            c_ft.fill = PatternFill("solid", fgColor="FAD7A0")
            c_ft.font = Font(bold=True, color=RED_HEX)

        # Clipboard
        _cell(ws_s, i, col, s.clipboard_events); col += 1

        # Texto
        _cell(ws_s, i, col, s.text_transfers); col += 1

        # Elevación
        c_elev = _cell(ws_s, i, col, "SÍ" if s.elevated else "No", bold=s.elevated); col += 1
        if s.elevated:
            c_elev.fill = PatternFill("solid", fgColor="FADBD8")
            c_elev.font = Font(bold=True, color=RED_HEX)

        # IPs
        _cell(ws_s, i, col, " | ".join(s.all_ips) if s.all_ips else "—"); col += 1

        # Riesgo
        _risk_cell(ws_s, i, col, s.risk); col += 1
        _cell(ws_s, i, col, s.risk_score); col += 1

        if has_incident:
            _proximity_cell(ws_s, i, col, s.incident_proximity_label); col += 1
        if has_country:
            cc = s.country_classification
            c_cc = _cell(ws_s, i, col, cc or "—"); col += 1
            if cc == "Sospechosa":
                c_cc.fill = PatternFill("solid", fgColor="FADBD8")
                c_cc.font = Font(bold=True, color=RED_HEX)
            elif cc == "Informativa":
                c_cc.fill = PatternFill("solid", fgColor="D5F5E3")

    widths_s = [5, 16, 18, 16, 20, 22, 22, 12, 14, 11, 10, 10, 22, 10, 7]
    if has_incident:
        widths_s.append(20)
    if has_country:
        widths_s.append(18)
    _set_widths(ws_s, widths_s)

    # ══════════════════════════════════════════════════════════════════════
    #  Hoja 3: Conexiones
    # ══════════════════════════════════════════════════════════════════════
    ws_c = wb.create_sheet("Conexiones")
    hdrs_c = ["#", "RMM", "Dirección", "Fecha/Hora", "Usuario", "ID Remoto", "Alias", "Hostname"]
    for c, h in enumerate(hdrs_c, 1):
        _hdr(ws_c, 1, c, h)

    for i, conn in enumerate(connections, 2):
        _cell(ws_c, i, 1, i - 1)
        _rmm_cell(ws_c, i, 2, conn.rmm_type.value)
        dir_val = conn.direction.value if isinstance(conn.direction, ConnectionDirection) else str(conn.direction)
        c3 = _cell(ws_c, i, 3, dir_val, bold=True)
        if dir_val == "Incoming":
            c3.font = Font(bold=True, color=GREEN_HEX)
        else:
            c3.font = Font(bold=True, color="1A5C87")
        _cell(ws_c, i, 4, fmt_dt(conn.datetime) or conn.dt_str)
        _cell(ws_c, i, 5, conn.user)
        _cell(ws_c, i, 6, conn.remote_id)
        _cell(ws_c, i, 7, conn.alias)
        _cell(ws_c, i, 8, conn.hostname)

    _set_widths(ws_c, [5, 16, 14, 22, 18, 16, 22, 18])

    # ══════════════════════════════════════════════════════════════════════
    #  Hoja 4: Indicadores_Riesgo
    # ══════════════════════════════════════════════════════════════════════
    ws_r = wb.create_sheet("Indicadores_Riesgo")
    hdrs_r = ["Sesión#", "RMM", "ID Remoto", "Inicio", "Nivel", "Score", "Razones"]
    for c, h in enumerate(hdrs_r, 1):
        _hdr(ws_r, 1, c, h)

    for i, s in enumerate(sessions, 2):
        _cell(ws_r, i, 1, i - 1)
        _rmm_cell(ws_r, i, 2, s.rmm_type.value)
        _cell(ws_r, i, 3, s.remote_id)
        _cell(ws_r, i, 4, fmt_dt(s.start_dt))
        _risk_cell(ws_r, i, 5, s.risk)
        _cell(ws_r, i, 6, s.risk_score)
        _cell(ws_r, i, 7, "\n".join(f"• {rz}" for rz in s.risk_reasons) if s.risk_reasons else "—")

    _set_widths(ws_r, [8, 16, 16, 22, 10, 7, 80])
    ws_r.row_dimensions[1].height = 20

    # ══════════════════════════════════════════════════════════════════════
    #  Hoja 5: IPs
    # ══════════════════════════════════════════════════════════════════════
    ws_ip = wb.create_sheet("IPs")
    hdrs_ip = [
        "IP", "Tipo", "País", "ISP",
        "AbuseIPDB Score", "Reportes", "Uso",
        "TOR",
        "VT Maliciosos", "VT Sospechosos", "VT Reputación",
        "CriminalIP Risk", "CriminalIP Score",
        "VPN", "Proxy",
    ]
    for c, h in enumerate(hdrs_ip, 1):
        _hdr(ws_ip, 1, c, h)

    for r, e in enumerate(ip_results, 2):
        _cell(ws_ip, r, 1, e.ip)
        _cell(ws_ip, r, 2, e.ip_type)
        _cell(ws_ip, r, 3, e.country)
        _cell(ws_ip, r, 4, e.isp)

        # AbuseIPDB Score
        sc = e.abuse_score
        c5 = _cell(ws_ip, r, 5, sc if sc is not None else "")
        if isinstance(sc, (int, float)) and sc is not None:
            if sc >= 75:
                c5.fill = PatternFill("solid", fgColor="FADBD8")
                c5.font = Font(bold=True, color=RED_HEX)
            elif sc >= 25:
                c5.fill = PatternFill("solid", fgColor="FDEBD0")
                c5.font = Font(bold=True, color=ORANGE_HEX)
            else:
                c5.fill = PatternFill("solid", fgColor="D5F5E3")

        # Reportes
        _cell(ws_ip, r, 6, e.abuse_reports if e.abuse_reports is not None else "")
        # Uso
        _cell(ws_ip, r, 7, e.abuse_usage)

        # TOR (combined)
        is_tor = e.is_tor
        c8 = _cell(ws_ip, r, 8, "SÍ" if is_tor else "No")
        if is_tor:
            c8.fill = PatternFill("solid", fgColor="FADBD8")
            c8.font = Font(bold=True, color=RED_HEX)

        # VirusTotal
        vt_m = e.vt_malicious
        c9 = _cell(ws_ip, r, 9, vt_m if vt_m is not None else "")
        if isinstance(vt_m, int) and vt_m > 0:
            c9.fill = PatternFill("solid", fgColor="FADBD8")
            c9.font = Font(bold=True, color=RED_HEX)
        _cell(ws_ip, r, 10, e.vt_suspicious if e.vt_suspicious is not None else "")
        _cell(ws_ip, r, 11, e.vt_reputation if e.vt_reputation is not None else "")

        # CriminalIP Risk
        cip_risk = e.criminalip_risk
        c12 = _cell(ws_ip, r, 12, cip_risk or "—")
        if cip_risk in ("dangerous", "critical"):
            c12.fill = PatternFill("solid", fgColor="FADBD8")
            c12.font = Font(bold=True, color=RED_HEX)
        elif cip_risk == "moderate":
            c12.fill = PatternFill("solid", fgColor="FDEBD0")
            c12.font = Font(bold=True, color=ORANGE_HEX)
        elif cip_risk in ("low", "safe"):
            c12.fill = PatternFill("solid", fgColor="D5F5E3")

        # CriminalIP Score
        _cell(ws_ip, r, 13, e.criminalip_score if e.criminalip_score is not None else "")

        # VPN
        c14 = _cell(ws_ip, r, 14, "SÍ" if e.criminalip_is_vpn else "No")
        if e.criminalip_is_vpn:
            c14.fill = PatternFill("solid", fgColor="FDEBD0")

        # Proxy
        c15 = _cell(ws_ip, r, 15, "SÍ" if e.criminalip_is_proxy else "No")
        if e.criminalip_is_proxy:
            c15.fill = PatternFill("solid", fgColor="FDEBD0")

    _set_widths(ws_ip, [17, 12, 8, 30, 15, 10, 22, 10, 14, 14, 12, 14, 14, 10, 10])

    # ══════════════════════════════════════════════════════════════════════
    #  Hoja 6: Correlación_Cross_RMM  (solo si cross_rmm_ips tiene datos)
    # ══════════════════════════════════════════════════════════════════════
    cross_ips = {ip: rmms for ip, rmms in summary.cross_rmm_ips.items()
                 if len(rmms) > 1}
    if cross_ips:
        ws_xr = wb.create_sheet("Correlación_Cross_RMM")
        hdrs_xr = ["IP", "RMMs", "País", "ISP", "Abuse Score", "VT Malicious"]
        for c, h in enumerate(hdrs_xr, 1):
            _hdr(ws_xr, 1, c, h)

        sorted_cross = sorted(cross_ips.items(), key=lambda x: (-len(x[1]), x[0]))

        for r, (ip, rmms) in enumerate(sorted_cross, 2):
            e = enriched_map.get(ip)
            _cell(ws_xr, r, 1, ip)
            c2 = _cell(ws_xr, r, 2, ", ".join(sorted(rmms)), bold=True)
            if len(rmms) >= 3:
                c2.fill = PatternFill("solid", fgColor="FADBD8")
                c2.font = Font(bold=True, color=RED_HEX)
            elif len(rmms) >= 2:
                c2.fill = PatternFill("solid", fgColor="FDEBD0")
                c2.font = Font(bold=True, color=ORANGE_HEX)
            _cell(ws_xr, r, 3, e.country if e else "")
            _cell(ws_xr, r, 4, e.isp if e else "")
            abuse_sc = e.abuse_score if e and e.abuse_score is not None else ""
            c5 = _cell(ws_xr, r, 5, abuse_sc)
            if isinstance(abuse_sc, (int, float)):
                if abuse_sc >= 75:
                    c5.fill = PatternFill("solid", fgColor="FADBD8")
                    c5.font = Font(bold=True, color=RED_HEX)
                elif abuse_sc >= 25:
                    c5.fill = PatternFill("solid", fgColor="FDEBD0")
                    c5.font = Font(bold=True, color=ORANGE_HEX)
            vt_m = e.vt_malicious if e else None
            c6 = _cell(ws_xr, r, 6, vt_m if vt_m is not None else "")
            if isinstance(vt_m, int) and vt_m > 0:
                c6.fill = PatternFill("solid", fgColor="FADBD8")
                c6.font = Font(bold=True, color=RED_HEX)

        _set_widths(ws_xr, [17, 40, 8, 30, 15, 14])

    # ══════════════════════════════════════════════════════════════════════
    #  Hoja 7: Proximidad_Incidente  (solo si hay fecha de incidente)
    # ══════════════════════════════════════════════════════════════════════
    if has_incident:
        ws_pi = wb.create_sheet("Proximidad_Incidente")
        hdrs_pi = [
            "Sesión#", "RMM", "ID Remoto", "Inicio",
            "Proximidad", "Horas al incidente",
            "Clasificación País", "Riesgo",
        ]
        for c, h in enumerate(hdrs_pi, 1):
            _hdr(ws_pi, 1, c, h)

        # Build categorized lists
        groups = [
            ("CRÍTICO (±24h)", summary.sessions_within_24h, "FDE8E8"),
            ("ALTO (±3d)",     summary.sessions_within_3d,  "FEF5E4"),
            ("MEDIO (±7d)",    summary.sessions_within_7d,  "FEFAE4"),
        ]

        r = 2
        for window_label, sess_list, row_bg in groups:
            if not sess_list:
                continue
            for s in sess_list:
                _cell(ws_pi, r, 1, s.idx, bg=row_bg)
                _rmm_cell(ws_pi, r, 2, s.rmm_type.value)
                _cell(ws_pi, r, 3, s.remote_id)
                _cell(ws_pi, r, 4, fmt_dt(s.start_dt))
                _proximity_cell(ws_pi, r, 5, window_label)
                hrs = s.incident_proximity_hours
                _cell(ws_pi, r, 6, f"{hrs:.1f}" if hrs is not None else "—")
                cc = s.country_classification
                c7 = _cell(ws_pi, r, 7, cc or "—")
                if cc == "Sospechosa":
                    c7.fill = PatternFill("solid", fgColor="FADBD8")
                    c7.font = Font(bold=True, color=RED_HEX)
                elif cc == "Informativa":
                    c7.fill = PatternFill("solid", fgColor="D5F5E3")
                _risk_cell(ws_pi, r, 8, s.risk)
                r += 1

        # Patrones anómalos
        if summary.anomalous_patterns:
            r += 1
            _hdr(ws_pi, r, 1, "PATRONES ANÓMALOS DETECTADOS", bg="0F5555", fg=WHITE)
            ws_pi.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(hdrs_pi))
            r += 1
            for pattern in summary.anomalous_patterns:
                _cell(ws_pi, r, 1, "⚠")
                _cell(ws_pi, r, 2, pattern, bold=True, fg=ORANGE_HEX)
                ws_pi.merge_cells(start_row=r, start_column=2, end_row=r, end_column=len(hdrs_pi))
                r += 1

        _set_widths(ws_pi, [8, 16, 16, 22, 20, 18, 18, 10])

    # ══════════════════════════════════════════════════════════════════════
    #  Hoja 8: Timeline
    # ══════════════════════════════════════════════════════════════════════
    ws_tl = wb.create_sheet("Timeline")
    hdrs_tl = ["Fecha/Hora", "RMM", "Tipo Evento", "ID Remoto", "Alias", "Detalle", "Riesgo"]
    for c, h in enumerate(hdrs_tl, 1):
        _hdr(ws_tl, 1, c, h)

    tl: list[tuple] = []
    for s in sessions:
        rmm_name = s.rmm_type.value
        if s.start_dt:
            detail_parts = []
            if s.file_transfers:
                detail_parts.append(f"Transferencias: {s.file_transfers}")
            if s.clipboard_events:
                detail_parts.append(f"Clipboard: {s.clipboard_events}")
            if s.text_transfers:
                detail_parts.append(f"Texto: {s.text_transfers}")
            if s.elevated:
                detail_parts.append("Elevado")
            detail = ", ".join(detail_parts) if detail_parts else "—"
            tl.append((s.start_dt, rmm_name, "Sesión iniciada",
                        s.remote_id, s.alias, detail, s.risk))
        if s.end_dt:
            tl.append((s.end_dt, rmm_name, "Sesión cerrada",
                        s.remote_id, s.alias,
                        f"Duración: {fmt_dur(s.duration_sec)}", s.risk))

    for conn in connections:
        rmm_name = conn.rmm_type.value
        dt_ = conn.datetime
        if dt_:
            dir_val = conn.direction.value if isinstance(conn.direction, ConnectionDirection) else str(conn.direction)
            detail = f"Usuario: {conn.user}" if conn.user else ""
            tl.append((dt_, rmm_name, f"Conexión {dir_val}",
                        conn.remote_id, conn.alias, detail, ""))

    tl.sort(key=lambda x: x[0] or datetime.min)

    for r, row in enumerate(tl, 2):
        dt_, rmm_n, ev, rid, alias, det, risk_ = row
        _cell(ws_tl, r, 1, fmt_dt(dt_))
        _rmm_cell(ws_tl, r, 2, rmm_n)
        _cell(ws_tl, r, 3, ev)
        _cell(ws_tl, r, 4, rid)
        _cell(ws_tl, r, 5, alias)
        _cell(ws_tl, r, 6, det)
        if risk_:
            _risk_cell(ws_tl, r, 7, risk_)
        else:
            _cell(ws_tl, r, 7, "—")

    _set_widths(ws_tl, [22, 16, 22, 16, 20, 55, 10])

    # ══════════════════════════════════════════════════════════════════════
    #  Ajustes finales
    # ══════════════════════════════════════════════════════════════════════
    for sh in wb.worksheets:
        sh.freeze_panes = "A2"
        if sh.max_row > 1 and sh.max_column > 1:
            try:
                sh.auto_filter.ref = sh.dimensions
            except Exception:
                pass

    wb.save(outpath)
    cprint(f"  [+] XLSX guardado: {outpath}", Fore.GREEN)
